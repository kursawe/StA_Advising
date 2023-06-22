import argparse
import openpyxl
import os
import pandas as pd
import termcolor
import re
import collections

# os.system('color')

module_catalogue_location = os.path.join(os.path.dirname(__file__),'module_catalogue','Module_catalogue.xlsx') 
module_catalogue = pd.read_excel(module_catalogue_location)

def check_form_file(filename):
    """preforms all advising checks on the 
    submitted form.
    
    Parameters:
    -----------
    
    filename : string
        path to the file that is being investigated,
        i.e. a filled-in module choice form
    """ 
    student = parse_excel_form(filename)
    print('Processing student with ID')
    print(student.student_id)
    print(' ')
    
    missed_programme_requirements, programme_adviser_recommendations = find_missing_programme_requirements(student)
    
    print('The student is missing the following programme requirements:')
    colour_code_print_statement(missed_programme_requirements)

    missed_prerequisites, prerequisite_adviser_recommendations = find_missing_prerequisites(student)
    
    print('The student is missing the following prerequisites')
    colour_code_print_statement(missed_prerequisites)

    not_running_modules, scheduling_adviser_recommendations = find_not_running_modules(student)

    print('The student selected the following modules when they are not running')
    colour_code_print_statement(not_running_modules)

    timetable_clashes, timetable_adviser_recommendations = find_timetable_clashes(student)
    print('I found the following timetable clashes')
    colour_code_print_statement(timetable_clashes)

    adviser_recommendations = merge_list_to_long_string([programme_adviser_recommendations, prerequisite_adviser_recommendations, 
                                                         scheduling_adviser_recommendations, timetable_adviser_recommendations])

    print('I have the following comments to the adviser:')
    colour_code_print_statement(adviser_recommendations, is_advice = True)

    summary_data = [student.student_id, missed_programme_requirements, missed_prerequisites, not_running_modules, adviser_recommendations]
    summary_data_frame = pd.DataFrame([summary_data], columns = ['Student ID', 'Unmet programme requirements', 'Missing prerequisites', 'Modules not running', 'Adviser recommendations'])

    return summary_data_frame

def find_timetable_clashes(student):
    """Find all modules that the student is planning to take and which are not actually running in the year and semester
    they are claiming.
    
    Parameters:
    -----------
    
    student : instance of Student class
        the student we are investigating
        
    Returns:
    --------
    
    timetable_clashes : string
        comments on all scheduled timetable clashes.
        
    adviser_recommendations : string
        advising recommendations, in this case may include warnings where scheduling is not finalised, or when we couldn't parse all scheduled events.
    """
    # make a list of timetable clashes
    timetable_clashes_list = []
    adviser_recommendations_list = []
    
    # get remaining honours years
    remaining_honours_years = student.honours_module_choices['Honours year'].unique()
    for honours_year in remaining_honours_years:
        for semester in ['S1', 'S2']:
            semester_modules = student.honours_module_choices[(student.honours_module_choices['Semester'] == semester) &  
                                                               (student.honours_module_choices['Honours year'] == honours_year)]['Module code'].to_list()
            timeslot_dictionary = dict()
            for module in semester_modules:
                these_timeslots = get_timeslots_for_module(module)
                timeslot_dictionary[module] = these_timeslots
            
            timetable_clashes_list += find_clashing_timeslots_and_modules(timeslot_dictionary, honours_year, semester)
            reduced_timeslot_dictionary = timeslot_dictionary.copy()
            for _,value in reduced_timeslot_dictionary.items():
                for entry in value:
                    entry.replace(' (even weeks)','')
            timetable_clashes_list += find_clashing_timeslots_and_modules(timeslot_dictionary, honours_year, semester)

            reduced_timeslot_dictionary = timeslot_dictionary.copy()
            for _,value in reduced_timeslot_dictionary.items():
                for entry in value:
                    entry.replace(' (odd weeks)','')
            timetable_clashes_list += find_clashing_timeslots_and_modules(timeslot_dictionary, honours_year, semester)

    # merge all found problems into a string
    timetable_clashes = merge_list_to_long_string(timetable_clashes_list)
    adviser_recommendations = merge_list_to_long_string(adviser_recommendations_list)
    
    return timetable_clashes, adviser_recommendations
 
def find_clashing_timeslots_and_modules(module_dictionary, honours_year, semester):
    """Given timeslot and a dictionary of concurrently running modules return the timeslots that are clashing and the clashing module codes
    
    Parameters:
    -----------
    
    module_dictionary : dictionary
        keys are module codes, values are lists of strings, which represent timeslots
    
    honours_year : string
        the honours year
        
    semester : string
        the semester
    
    Returns:
    --------

    timetable_clashes_list : list of strings
        warning messages about clashing modules
    """
    timetable_clashes_list = []
    all_timeslots = []
    for module, timeslots in module_dictionary.items():
        all_timeslots += timeslots

    my_timeslot_counter = collections.Counter(all_timeslots)
    duplicate_entries = [timeslot for timeslot, count in my_timeslot_counter.items() if count > 1]
    for timeslot in duplicate_entries:
        clashing_module_codes = []
        for module, timeslot_list in module_dictionary.items():
            if timeslot in timeslot_list:
                clashing_module_codes.append(module)
        warning_string = 'Found timeslot clash for ' + honours_year + ' ' + semester + ' at ' + timeslot + ' between modules '
        for module_index, module in enumerate(clashing_module_codes):
            warning_string += module
            if module_index < len(clashing_module_codes) - 1:
                warning_string += ' and '
        timetable_clashes_list.append(warning_string)
    
    return timetable_clashes_list
   
def get_timeslots_for_module(module):
    """Returns all timeslots for a module
    
    Parameters:
    -----------
    
    module : string
        the module code we are interested in.
        
    Returns:
    --------
    
    timeslots : list of strings
        all timeslots that the module is running in
    """
    timeslot_entry = module_catalogue[module_catalogue['Module code'] == module]['Timetable'].values[0]
    timeslots = []
    if isinstance(timeslot_entry,str):
        timeslot_splits = timeslot_entry.split()
        first_timeslot = timeslot_splits[0]
        if timeslot_splits[2].startswith('('):
            second_timeslot = first_timeslot + ' ' + timeslot_splits[4][:-1]
            third_timeslot = first_timeslot + ' ' + timeslot_splits[5]
            if third_timeslot.endswith(','):
                third_timeslot = third_timeslot[:-1]
            first_timeslot += ' ' + timeslot_splits[1] + ' ' + timeslot_splits[2] + ' ' + timeslot_splits[3][:-1]
            timeslots += [first_timeslot, second_timeslot, third_timeslot]
            if len(timeslot_splits) > 7:
                remaining_splits = timeslot_splits[6:]
            else:
                remaining_splits = []
        else:
            first_timeslot += ' ' + timeslot_splits[1]
            if first_timeslot.endswith(','):
                first_timeslot = first_timeslot[:-1]
            if len(timeslot_splits) > 2:
                remaining_splits = timeslot_splits[2:]
            else:
                remaining_splits = []
            timeslots.append(first_timeslot)
        
        current_index = 0
        while current_index < len(remaining_splits):
            this_timeslot = remaining_splits[current_index] + ' ' +  remaining_splits[current_index + 1]
            if this_timeslot.endswith(','):
                this_timeslot = this_timeslot[:-1]
            timeslots.append(this_timeslot)
            current_index +=2
        
    return timeslots

def find_not_running_modules(student):
    """Find all modules that the student is planning to take and which are not actually running in the year and semester
    they are claiming.
    
    Parameters:
    -----------
    
    student : instance of Student class
        the student we are investigating
        
    Returns:
    --------
    
    not_running_modules : string
        comments on all modules that won't be running as intended.
        
    adviser_recommendations : string
        advising recommendations, in this case may include warnings where scheduling is not finalised.
    """
    # make a list of not running modules
    not_running_modules_list = []
    adviser_recommendations_list = []

    for _, row in student.honours_module_choices.iterrows():
        # get the module data
        planned_module_code = row['Module code']
        planned_academic_year = row['Academic year']
        planned_semester = row['Semester']
        module_catalogue_entry = module_catalogue[module_catalogue['Module code'] == planned_module_code]
        module_semester = module_catalogue_entry['Semester'].values[0]
        # tell if the student picked the wrong semester
        if planned_semester != module_semester:
            not_running_modules_list.append('Selected module ' + planned_module_code + ' for Semester ' +
                                            planned_semester + ' but it is actually running in ' + module_semester)
        # figure out when the module is running
        module_academic_year = module_catalogue_entry['Year'].values[0]
        module_is_alternating_entry = module_catalogue_entry['Alternate years'].values[0]
        if module_is_alternating_entry == 'Yes':
            module_is_alternating = True
        elif module_is_alternating_entry == 'No':
            module_is_alternating = False
        else:
            raise(ValueError('cannot tell if module ' + planned_module_code + ' is alternating or not. Check the table entry.'))
        # figure out which years the module is running in
        list_of_running_academic_years = [module_academic_year]
        start_year = int(module_academic_year[:4])
        for repeat_index in range(4):
            if module_is_alternating:
                new_academic_year = str(start_year + 2*repeat_index) + '/' + str(start_year + 2*repeat_index + 1)
            else:
                new_academic_year = str(start_year + repeat_index) + '/' + str(start_year + repeat_index + 1)
            list_of_running_academic_years.append(new_academic_year)
        if planned_academic_year not in list_of_running_academic_years:
            not_running_modules_list.append('Selected module ' + planned_module_code + ' is not running in academic year ' +
                                            str(planned_academic_year))

    # merge all found problems into a string
    not_running_modules = merge_list_to_long_string(not_running_modules_list)
    adviser_recommendations = merge_list_to_long_string(adviser_recommendations_list)
    
    return not_running_modules, adviser_recommendations


def find_missing_prerequisites(student):
    """find any missing prerequisites or violated anti-requisites.
    
    Parameters :
    -----------
    
    student : instance of Student class
        can be generated with 'parse_excel_form()'
        
    Returns :
    ---------

    missed_prerequisites : string
        Unmet programme requirements and violated anti-requisites. Will return 'None' if all prerequisites are met and
        no anti-requisite is violated.
        
    adviser_recommendations : string
        advising recommendations, in this case may include reminders to get letters of approval.
    """
    list_of_missed_prerequisites = []
    list_of_recommendations = []
    
    for module in student.planned_honours_modules:
        these_missing_prerequisites, these_adviser_recommendations = get_missing_prerequisites_for_module(module, student)
        list_of_missed_prerequisites += [these_missing_prerequisites]
        list_of_recommendations += [these_adviser_recommendations]
        
    # merge all missed prerequisites into a string
    missed_prerequisites = merge_list_to_long_string(list_of_missed_prerequisites)
    adviser_recommendations = merge_list_to_long_string(list_of_recommendations)

    return missed_prerequisites, adviser_recommendations
    
def get_missing_prerequisites_for_module(module, student):
    """Find which prerequisites the student is missing for the given module.
    Will also check anti-requisites.
    
    Parameters :
    ------------
    
    module : string
        module code for the module that we are investigating
        
    student : instance of Student class
        The student we are checking
        
    Returns :
    ---------

    missed_prerequisites : string
        missing prerequisites and violated anti-requisites
        
    adviser_recommendation : string
        any relevant adviser recommndations for this module
    """
    # make a list of missing prerequisites
    missed_prerequisites_list = []
    adviser_recommendations_list = []

    # find which year and semester the module is selected for
    year_of_this_module = student.honours_module_choices[student.honours_module_choices['Module code'] == module]['Honours year'].values[0]
    year_number_of_this_module = int(year_of_this_module[-1])
    semester_of_this_module = student.honours_module_choices[student.honours_module_choices['Module code'] == module]['Semester'].values[0]

    # construct a list of all courses the student has taken by then
    # and construct a list of all modules the student is taking concurrently
    previously_taken_modules = student.passed_modules.copy()
    simultaneously_taken_modules = []
    # we need this one for anti-requisites below
    modules_taken_in_same_year=[]

    for _, row in student.honours_module_choices.iterrows():
        year_number = int(row['Honours year'][-1])
        if year_number < year_number_of_this_module:
            previously_taken_modules.append(row['Module code'])
        if semester_of_this_module == 'S2':
            if row['Honours year'] == year_of_this_module and row['Semester'] == 'S1' :
                previously_taken_modules.append(row['Module code'])
        if (row['Honours year'] == year_of_this_module 
            and row['Semester'] == semester_of_this_module
            and row['Module code'] != module):
            simultaneously_taken_modules.append(row['Module code'])
        if (row['Honours year'] == year_of_this_module 
            and row['Module code'] != module):
            modules_taken_in_same_year.append(row['Module code'])

    # get pre-requisite string for that module
    prerequisites = module_catalogue[module_catalogue['Module code'] == module]['Prerequisites'].values[0]
    
    # if the prerequsites are not empty
    if  isinstance(prerequisites,str) and module!='MT5867':
        prerequisite_list = prerequisites.split()
        # if there is only one prerequisite we can easily parse that
        if len(prerequisite_list) == 1:
            # this must be a module code now
            required_module = prerequisite_list[0]
            if required_module not in previously_taken_modules:
                missed_prerequisites_list.append('Student is missing prerequisite ' + required_module + ' for module ' + module)
        # sometiems it's just a letter of agreement that we need to know about
        elif len(prerequisite_list) == 3:
            if prerequisites == 'Letter of agreement':
                adviser_recommendations_list.append('Module ' + module + ' requires a letter of agreement')
        else:
            #now there is a boolean statement coming, so we turn the module codes into boolean strings and evaluate the outcome
            # (thanks, ChatGPT)
            module_codes = re.findall(r'[A-Z]{2}\d{4}', prerequisites)
            parsed_prerequisites = prerequisites
            for module_code in module_codes:
                # co-requisites are preceded by the word co-requisite and don't have brackets after
                if module_code in prerequisite_list:
                    location_index = prerequisite_list.index(module_code)
                    if location_index > 0:
                        if prerequisite_list[location_index - 1] == 'co-requisite':
                            if module_code in previously_taken_modules or module_code in simultaneously_taken_modules:
                                parsed_prerequisites = parsed_prerequisites.replace(module_code, 'True')
                            else:
                                parsed_prerequisites = parsed_prerequisites.replace(module_code, 'False')
                            parsed_prerequisites = parsed_prerequisites.replace('co-requisite ', '')
                        else: 
                            # the location index is larger than one but we are not at a co-requisite
                            if module_code in previously_taken_modules:
                                parsed_prerequisites = parsed_prerequisites.replace(module_code, 'True')
                            else:
                                parsed_prerequisites = parsed_prerequisites.replace(module_code, 'False')
                    # the location index is zero and we just check the module code
                    if module_code in previously_taken_modules:
                        parsed_prerequisites = parsed_prerequisites.replace(module_code, 'True')
                    else:
                        parsed_prerequisites = parsed_prerequisites.replace(module_code, 'False')
                else:
                    if module_code in previously_taken_modules:
                        parsed_prerequisites = parsed_prerequisites.replace(module_code, 'True')
                    else:
                        parsed_prerequisites = parsed_prerequisites.replace(module_code, 'False')
            prerequisites_are_met = eval(parsed_prerequisites)
            if not prerequisites_are_met:
                missed_prerequisites_list.append('Student is missing prerequisite [' + prerequisites+ '] for module ' + module + 
                                                 ' ([' + parsed_prerequisites + '])')
            
    # MT5867 is a special case that I don't know how to parse automatically:
    if module == 'MT5867':
        prerequisites = 'two of (MT3505, MT4003, MT4004, MT4512, MT4514, MT4515, MT4526)'
        list_of_modules = ['MT3505', 'MT4003', 'MT4004', 'MT4512', 'MT4514', 'MT4515', 'MT4526']
        number_of_matching_modules = len(set.intersection(set(list_of_modules), set(previously_taken_modules)))
        if number_of_matching_modules <2:
            missed_prerequisites_list.append('Student is missing prerequisite [' + prerequisites + '] for module MT5867')
            
    # now check anti-requisites:
    # to do so, get the anti-requisites
    antirequisites = module_catalogue[module_catalogue['Module code'] == module]['Antirequisites'].values[0]
    if isinstance(antirequisites, str):
        # check any listed module code individually
        anti_module_codes = re.findall(r'[A-Z]{2}\d{4}', antirequisites)
        for module_code in anti_module_codes:
            if module_code in previously_taken_modules or module_code in simultaneously_taken_modules:
                missed_prerequisites_list.append('Student selected antirequisite ' + module_code + ' for module ' + module)

    
    # merge all missed prerequisites into a string
    missed_prerequisites = merge_list_to_long_string(missed_prerequisites_list)
    adviser_recommendations = merge_list_to_long_string(adviser_recommendations_list)

    return missed_prerequisites, adviser_recommendations

def colour_code_print_statement(print_statement, is_advice = False):
    """Prints the given string in red if the string is not equal to 'None'.
    Otherwise it will print in green.
    
    Parameters:
    -----------
    
    print_statement : string
        the statement to be printed
        
    is_advice : bool
        if False, strings that are not 'None' will be printed in light magenta instead of red
    """
    if print_statement != 'None':
        if is_advice:
            print(termcolor.colored(print_statement,'blue'))
        else:
            print(termcolor.colored(print_statement,'red'))
    else:
        print(termcolor.colored('None','green'))
        
    #add an empty line by default
    print(' ')

def find_missing_programme_requirements(student):
    """check that the student fulfils their honours requirements
    
    Parameters :
    -----------
    
    student : instance of Student class
        can be generated with 'parse_excel_form()'
        
    Returns :
    ---------

    missed_requirements : string
        Unmet programme requirements. Will return 'None' if all programme requirements are met
        
    adviser_recommendations : string
        advising recommendations that don't strictly count as unmet programme requirements
    """
    
    list_of_missed_requirements = []
    list_of_adviser_recommendations = []
    if student.programme_name == 'Bachelor of Science (Honours) Mathematics':
        # check the credit load
        missed_requirement, adviser_recommendation = check_for_120_credits_each_year(student)
        list_of_missed_requirements.append(missed_requirement)
        list_of_adviser_recommendations.append(adviser_recommendation)

        # check there are four modules in MT3501 to MT3508
        list_of_MT350X_modules = ['MT3501', 'MT3502', 'MT3503', 'MT3504', 'MT3505', 'MT3506', 'MT3507', 'MT3508']
        number_of_MT350X_modules = student.get_number_of_modules_in_list(list_of_MT350X_modules)
        if number_of_MT350X_modules < 4:
            list_of_missed_requirements.append('Student is only taking ' + str(number_of_MT350X_modules) + 'out of MT3501-MT3508')
            
        # check there is a computing module
        list_of_computing_modules = ['MT3510', 'MT4111', 'MT4112', 'MT4113']
        number_of_computing_modules = student.get_number_of_modules_in_list(list_of_computing_modules)
        if number_of_computing_modules ==0:
            list_of_missed_requirements.append('Student is not taking a computing module')
        
        # check there is a final year project
        list_of_project_codes = ['MT4598', 'MT4599']
        number_final_year_projects= student.get_number_of_modules_in_list(list_of_project_codes)
        if number_final_year_projects !=1:
            list_of_missed_requirements.append('Student is not taking an allowed final year project')
        else:
            # check that the student is actually taking it in year 4
            if 'MT4598' in student.full_module_list:
                this_year = student.honours_module_choices[student.honours_module_choices['Module code'] == 'MT4598']['Honours year'].iloc[0]
            else:
                this_year = student.honours_module_choices[student.honours_module_choices['Module code'] == 'MT4599']['Honours year'].iloc[0]
            if this_year != 'Year 2':
                list_of_missed_requirements.append('Student is not taking their final year project in their final year.')
        
        # check that unallowed modules are not taken
        unallowed_modules = ['MT4794', 'MT4795', 'MT4796', 'MT4797']
        number_of_unallowed_modules = student.get_number_of_modules_in_list(unallowed_modules)
        if number_of_unallowed_modules >0:
            list_of_missed_requirements.append('Student is taking a module in MT4794-MT4797')
            
        # check dip-down and dip-across: no more than two modules should be outside of MT3X to MT5X
        list_of_all_non_maths_modules = [module for module in student.all_honours_modules if 'MT2' not in module 
                                                                                       and 'MT3' not in module 
                                                                                        and 'MT4' not in module 
                                                                                        and 'MT5' not in module]
        if len(list_of_all_non_maths_modules) >2:
            list_of_missed_requirements.append('Student is taking more than 2 modules as dip-down or dip-across, which is not allowed.')

        # check that there are at least 90 credits (6 modules) at 4000 level or above
        list_of_4000_and_5000_modules = [module for module in student.all_honours_modules if 'MT4' in module or 'MT5' in module]
        if len(list_of_4000_and_5000_modules) <6:
            list_of_missed_requirements.append('Student is not planning to take enough credits at 4000 level or above')

        # remind advisers to get permissions
        list_of_planned_5000_level_modules = [module for module in student.planned_honours_modules if 'MT5' in module]
        if len(list_of_planned_5000_level_modules) >0:
            list_of_adviser_recommendations.append('Student is planning to take 5000 level modules (which will require permission)')

        list_of_2000_level_modules = [module for module in student.planned_honours_modules if 'MT2' in module]
        if len(list_of_2000_level_modules) >0:
            list_of_adviser_recommendations.append('Student is planning to take 2000 level modules, (which will require permission)')

        list_of_planned_non_maths_modules = [module for module in student.planned_honours_modules if 'MT2' not in module 
                                                                                        and 'MT3' not in module 
                                                                                        and 'MT4' not in module 
                                                                                        and 'MT5' not in module]
        if len(list_of_planned_non_maths_modules) >0:
            list_of_adviser_recommendations.append('Student is planning to take non-MT modules, which requires permission')

    # merge all missed requirements into a string
    missed_requirements = merge_list_to_long_string(list_of_missed_requirements)
    adviser_recommendations = merge_list_to_long_string(list_of_adviser_recommendations)

    return missed_requirements, adviser_recommendations
    
def check_for_120_credits_each_year(student):
    """check whether the student is actually taking 120 credits each acacemic year, and whether they 
       have an even split of modules.

    Parameters:
    -----------

    student : instance of Student()
        the student we are investigating

    Returns:
    --------

    missed_requirement : string
        a note if credits are not satisfied

    adviser_recommendation : string
        a note if module split is uneven 
    """
    honours_years = student.honours_module_choices['Honours year'].unique()
    list_of_missed_requirements = []
    list_of_adviser_recommendations = []
    
    #checking total number of modules
    for honours_year in honours_years:
        this_data_base = student.honours_module_choices[student.honours_module_choices['Honours year'] == honours_year]
        if honours_year == 'Year 1' or honours_year == 'Year 2':
            if len(this_data_base)!=8:
                list_of_missed_requirements.append('Not collecting 120 credits in ' + honours_year)
        if honours_year == 'Year 3' and len(this_data_base)!=7:
            list_of_missed_requirements.append('Not collecting 120 credits in ' + honours_year)
    
    #checking moduel splits
    for honours_year in honours_years:
        if honours_year == 'Year 1' or (honours_year == 'Year 2' and student.expected_honours_years == 3):
            for semester in ['S1', 'S2']:
                this_smaller_data_base = this_data_base[this_data_base['Semester'] == semester]
                if len(this_smaller_data_base) !=4:
                    list_of_adviser_recommendations.append('Not taking even credit split in ' + honours_year)
        elif honours_year == 'Year 2':
            this_reduced_data_base = this_data_base[this_data_base['Module code'] != 'MT4599']
            semester_1_modules = this_reduced_data_base[this_reduced_data_base['Semester'] == 'S1']['Module code']
            semester_2_modules = this_reduced_data_base[this_reduced_data_base['Semester'] == 'S1']['Module code']
            if len(semester_1_modules) != 4 or len(semester_2_modules) != 3:
                list_of_adviser_recommendations.append('Student is taking a high course load in second semester of final honours year')
        elif honours_year == 'Year 3':
            this_reduced_data_base = this_data_base[this_data_base['Module code'] != 'MT5599']
            semester_1_modules = this_reduced_data_base[this_reduced_data_base['Semester'] == 'S1']['Module code']
            semester_2_modules = this_reduced_data_base[this_reduced_data_base['Semester'] == 'S1']['Module code']
            if len(semester_1_modules) != 3 or len(semester_2_modules) != 3:
                list_of_adviser_recommendations.append('Student is taking uneven course load in final honours year')
    
    missed_requirement = merge_list_to_long_string(list_of_missed_requirements)
    adviser_recommendation = merge_list_to_long_string(list_of_adviser_recommendations)
    
    return missed_requirement, adviser_recommendation

def merge_list_to_long_string(a_list):
    """takes a list of strings and returns a string that separates the entries with a comma and a space.
        Returns the string 'None' if the list is empty. If the first entry of the list is 'None' then that entry will be ignored.
    
    Parameters:
    -----------
    a_list : list
        the list we want to parse, for example a list of unmet programme requirements
    
    Returns:
    --------
    a_string : string
        contains all entries in a_list separated by a comma and a space
        is 'None' if a_list is empty
    """
    a_string = ''
    first_item_found = False
    for item in a_list:
        if item != 'None':
            if first_item_found:
                a_string += ', ' + item
            else:
                a_string = item
                first_item_found = False
    
    if a_string == '':
        a_string = 'None'
    
    return a_string
    
def parse_excel_form(filename):
    """returns an instance of a 'student' class
    that has all the excel data as named attributes

    Parameters:
    -----------
    
    filename : string
        path to the file that is being investigated,
        i.e. a filled-in module choice form
        
    Returns:
    --------

    student : instance of Student class
        an object with student attributes.
    """
    # open the form file and read in the student ID
    this_workbook = openpyxl.load_workbook(filename=filename)
    sheet = this_workbook.active
    student_id = sheet["D5"].value
    
    # Now that we have the student ID we can look up the student in the database:
    current_directory = os.path.dirname(os.path.abspath(__file__))
    potential_data_files = os.listdir(os.path.join(current_directory, 'student_data'))
    data_files = []
    for candidate_filename in potential_data_files:
       this_filename, file_extension = os.path.splitext(candidate_filename)
       if file_extension == '.csv':
           data_files.append(candidate_filename)
    
    # turn them all into data base files
    data_bases = []
    for data_file_name in data_files:
        data_path = os.path.join(current_directory, 'student_data', data_file_name)
        this_data_frame = pd.read_csv(data_path)
        data_bases.append(this_data_frame)
    
    # get a table with only the entries for this student
    student_not_yet_found = True

    data_base_index = 0
    while student_not_yet_found:
        this_data_base = data_bases[data_base_index]
        if student_id in this_data_base['Student ID'].to_numpy():
            student_not_yet_found = False
            student_data_base = this_data_base[this_data_base['Student ID'] == student_id]
        data_base_index += 1

    # infer the year of study from the earliest module taken
    data_of_module_years = student_data_base['Year'].str.slice(0,4).astype('int')
    earliest_year = data_of_module_years.min()
    
    #students who took their first module in 2021 will now be in year 3, i.e. 2023-2021+1
    year_of_study = 2023 - earliest_year + 1
    year_of_study = year_of_study
    
    # identify all modules that the student has passed
    data_base_of_passed_modules = student_data_base[student_data_base['Assessment result']=='P']
    passed_modules = data_base_of_passed_modules['Module code'].to_list()
    passed_modules = passed_modules
    
    #identify the programme of the student
    programme_entries = student_data_base['Programme name'].unique()    
    assert(len(programme_entries == 1))
    programme_name = programme_entries[0]
    
    # Figure out what year they are in and how many they have left
    if 'Bachelor of Science' in programme_name:
        no_of_programme_years = 4
        expected_honours_years = 2
    elif 'Master in Mathematics' in programme_name:
        no_of_programme_years = 5
        expected_honours_years = 3
    else: 
        print('what the heck kind of programme is that?? ' + programme_name)
    
    if 'EXA120' in student_data_base['Module code']:
        no_of_programme_years -=1
        
    no_subhonours_years = no_of_programme_years - expected_honours_years
    current_honours_year = year_of_study - no_subhonours_years
    
    # make a separate data base of passed honours modules
    passed_honours_modules = list()
    for previous_honours_year in range(1,current_honours_year):
        year_difference = current_honours_year - previous_honours_year  
        year_number = 23-year_difference
        calendar_year_string = '20' + str(calendar_year) + '/' + str(calendar_year + 1)
        data_base_of_passed_modules_this_year = data_base_of_passed_modules[data_base_of_passed_modules['year']==calendar_year_string]
        passed_modules_this_hear = data_base_of_passed_modules_this_year['Module code'].to_list()
        passed_honours_modules += passed_modules_this_hear
    
    # read in modules for all honours years that have not happened yet
    module_table = []
    
    for remaining_honours_year in range(current_honours_year,expected_honours_years + 1):
        year_key = 'Year ' + str(remaining_honours_year)
        calendar_year = 22 + remaining_honours_year
        calendar_year_string = '20' + str(calendar_year) + '/20' + str(calendar_year + 1)
        for semester_number in [1,2]:
            semester_modules = get_modules_under_header(sheet, year_key + ' of Honours: Semester ' + str(semester_number)) 
            for module in semester_modules:
                module_table.append([year_key, calendar_year_string, 'S' + str(semester_number), module,])

    # Turn this all into a nice pandas data frame
    honours_module_choices = pd.DataFrame(module_table, columns = ['Honours year', 'Academic year', 'Semester', 'Module code'])
    
    this_student = Student(student_id, 
                           programme_name, 
                           year_of_study,
                           expected_honours_years,
                           current_honours_year,
                           passed_modules,
                           passed_honours_modules,
                           honours_module_choices)
    
    # return the student
    return this_student

class Student():
    def __init__(self, 
                 student_id, 
                 programme_name,
                 year_of_study,
                 expected_honours_years,
                 current_honours_year,
                 passed_modules,
                 passed_honours_modules,
                 honours_module_choices):
        """Constructor for the Student class.
        
        Parameters:
        -----------
        
        student_id : int
            The student ID
            
        programme_name : string
            the full name of the programme, as it appears in the MMS databases
            
        year_of_study : int
            the current year of study
            
        expected_honours_years : int
            the number of expected honours years. This is slightly superfluous as it can be
            inferred from the programme, i.e. it will be 3 for masters programmes and 2 for
            bachelors programmes
            
        current_honours_year : int
            Which year of honours the student is in - this is not superfluous, as it cannot always
            be inferred from the year of study, since Advance Standing Credits (direct entry) need
            to be taken into account
            
        passed_modules : list of strings
            A list containing the module codes of all passed modules
            
        passed_honours_modules : list of strings
            A list of all honours modules that the student has already taken and passed
            
        honours_module_choices : Pandas data frame
            The honours module codes that the studen thas selected. The data frame current contains
            the following columns ['Honours year', 'Academic year', 'Semester', 'Module code']
        """
        
        self.student_id = student_id
        self.programme_name = programme_name
        self.year_of_study = year_of_study
        self.expected_honours_years = expected_honours_years
        self.current_honours_year = current_honours_year
        self.passed_modules = passed_modules
        self.passed_honours_modules = passed_honours_modules
        self.honours_module_choices = honours_module_choices

        # for convenience also make a list of all selected and taken modules
        self.full_module_list = self.passed_modules.copy()
        self.full_module_list += self.honours_module_choices['Module code'].to_list()
        
        # and a list of all selected and taken honours modules
        self.all_honours_modules = self.passed_honours_modules
        self.all_honours_modules += self.honours_module_choices['Module code'].tolist()
        
        # and a list of planned honours modules
        self.planned_honours_modules = self.honours_module_choices['Module code'].tolist()

    def get_number_of_modules_in_list(self, module_list):
        """Get the number of modules that the student is taking in the module list. Already passed modules and scheduled
        modules are both counted.
        
        Parameters:
        -----------
        
        module_list : list of strings
            list of module codes
            
        Returns: 
        --------
        
        number_of_modules : int
            the number of modules in the given list that the student is taking.
        """
        number_of_modules = len(set.intersection(set(self.full_module_list),set(module_list)))
        
        return number_of_modules
    
def get_modules_under_header(sheet, header):
    """get all the modules in the student module choice form under a given heading
    
    Parameters:
    -----------
    
    sheet : openpyxl sheet object
        generated from the workbook excel file
        
    header : string
        the header that we are investigating

    Returns:
    --------

    modules : list of strings
        module codes under the given header
    """
    modules = []
    for row in sheet.iter_rows():
        if isinstance(row[1].value, str):
            if header in row[1].value:
                row_number = row[1].row

    # the row number where modules are entered is 2 down
    row_number += 2
    next_cell_name_with_module = 'B' + str(row_number)
    module_code_is_not_empty = sheet[next_cell_name_with_module].value is not None

    while module_code_is_not_empty:
        modules.append(sheet[next_cell_name_with_module].value)
        row_number+=1 
        next_cell_name_with_module = 'B' + str(row_number)
        module_code_is_not_empty = sheet[next_cell_name_with_module].value is not None

    return modules

# def check_programme_requirements(Student)
#     """This checks if a student passes their programme requirements.
#     
#     Parameters:
#     -----------
#     
#     student : instance of Student class
#         The student we are investigating
#     """
#     # Look up math requirements as a list of conditions
#     # loop through the list of conditions
#     # check if conditions are fulfilled
#     
# def check_prerequisites(student):
#     """This checks if a student meets the prerequisits for their courses
#     
#     Parameters:
#     -----------
#     
#     student : instance of Student class
#         The student we are investigating
#     """
#     for module in get_all_student_modules(student):
#         check_specific_module_requirements(student, module)
# 
# def check_specific_module_requirements(module, student):
#     """This checks if a student meets the prerequisits for their courses
#     
#     Parameters:
#     -----------
#     
#     module : string
#         module code we are investigating
# 
#     student : instance of Student class
#         The student we are investigating
#     """
#     these_module_requirements = get_requirements_for_module(module_code)
#     previous_modules = get_previous_modules(student, module)
#     concurrent_modules = get_concurrent_modules(student, module)
#     for requirement in these_module_requirements:
#         if not requirement.is_fulfilled(student, previous_modules, concurrent_modules)
#             print('missing requirement ' + requirement.blurb())
#             
# def check_for_timetable_clashes(student):
#     """This checks if a student will have any timetable clashes within maths
#     
#     Parameters:
#     -----------
#     
#     student : instance of Student class
#         The student we are investigating
#     """
#     for year in student.get_honours_years():
#         for semester in [1,2]:
#             these_modules = student.get_modules(year, semester)
#             check_for_timetable_clashes(these_modules)

def colour_code_passes(column):
    passed_colour = 'background-color: palegreen;'
    failed_colour = 'background-color: lightcoral;'
    return [passed_colour if value=='None' else failed_colour for value in column]

def colour_recommendations(column):
    recommendation_colour = 'background-color: orange;'
    default_colour = ''
    return [default_colour if value=='None' else recommendation_colour for value in column]

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
                    prog='AdvisingScript',
                    description='This script helps with advising',
                    epilog='All results are experimental and not to be trusted. Double-check me. \n\n')
    
    parser.add_argument('filename')       
    args = parser.parse_args()
    form_filename = args.filename
    summary_data_frame = check_form_file(form_filename)
    summary_data_frame = (summary_data_frame.style.apply(colour_code_passes, subset = ['Unmet programme requirements', 'Missing prerequisites', 'Modules not running'], axis = 0).
                          apply(colour_recommendations, subset = ['Adviser recommendations'], axis = 0))

    saving_name = 'summary_file.xlsx'
    writer = pd.ExcelWriter(saving_name) 
    # Manually adjust the width of the last column
    summary_data_frame.to_excel(writer)
    # writer.book.close()
    worksheet = writer.sheets['Sheet1']
    # cell_format = writer.book.add_format({'font_size': '14'})
    font = openpyxl.styles.Font(size=14)
    worksheet.set_column(0,0,width=5)
    worksheet.set_column(1,1,width=20)
    worksheet.set_column(2,2,width=40)
    worksheet.set_column(3,3,width=40)
    worksheet.set_column(4,4,width=40)
    worksheet.set_column(5,5,width=40)
    # for column in 'ABCDEFG':
        # col = worksheet.column_dimensions[column]
        # column.font = font
    # for row in worksheet.iter_cols():
        # for cell in row:
            # cell.font = font
    writer.save()
    
    new_workbook = openpyxl.load_workbook(saving_name)

    # Access the active sheet
    worksheet = new_workbook.active

    # Set the font size for all cells in the worksheet
    font_size = 14
    font = openpyxl.styles.Font(size=font_size)
    for row in worksheet.iter_rows():
        for cell in row:
            cell.font = font

    # Make the first row bold
    bold_font = openpyxl.styles.Font(size=font_size, bold=True)
    for cell in worksheet[1]:
        cell.font = bold_font
    # Save the modified workbook
    new_workbook.save(saving_name)
   # parse command line
   # get list of form files
   # for each form file check student
   # check_form_file